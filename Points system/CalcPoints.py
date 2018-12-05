import openpyxl
import os
import math
# import tkinter

def calc_t(s):
    if s['A2'].value[8]=='a':
        col='M' #ao5s stored in column M
    elif s['A2'].value[8]=='m':
        col='J' #mo3s in column J
    elif s['A2'].value[-1]=='3':
        col='H' #bo3s in column H
    elif s['A2'].value[-1]=='2':
        col='G' #bo2s in column G
    else:
        col='E' #single solve in col E
    if s.title[:6]=='333mbf':
        if s['A2'].value[-1]=='3':
            col='Q' #mbld bo3s in column Q
        elif s['A2'].value[-1]=='2':
            col='M' #mbld bo2s in column M
        else:
            col='H' #mbld  single solve in col H
        for i in range(5,s.max_row+1):
            try:
                s[col+str(i)].value=1/(99-s[col+str(i)].value//10000000)  #get actual points (reciprocated, because in multi more is better)
            except:
                s[col+str(i)].value=None
            if s[col+str(i)].value==1/100:
                s[col+str(i)].value=None
    for i in range(5,s.max_row+1):  #calc t
        try:
            s['S'+str(i)].value=s[col+str(i)].value/s[col+'5'].value
        except:
            try: #to change from mm:ss:uu to seconds
                s['S'+str(i)].value=(s[col+str(i)].value.minute*60+s[col+str(i)].value.second+s[col+str(i)].value.microsecond/1000000)/(s[col+'5'].value.minute*60+s[col+'5'].value.second+s[col+'5'].value.microsecond/1000000)
            except:
                break
            
def calc_yp(s,n):
    i=s.max_row
    while s['S'+str(i)].value==None:
        i=i-1
    if s.title==n.title:
        s['T'+str(i)].value=(s.max_row-3-s['A'+str(i)].value)/(s.max_row-4)
        s['U'+str(i)].value=s['T'+str(i)].value/s['S'+str(i)].value
    else:
        j=n.max_row
        while n['S'+str(j)].value==None:
            j=j-1
        m=j
        while n['S'+str(j)].value>s['S'+str(i)].value:
            j=j-1
        if n['T'+str(j+1)].value==None:
            s['T'+str(i)].value=n['T'+str(j)].value
            s['U'+str(i)].value=s['T'+str(i)].value/s['S'+str(i)].value
        else:
            s['T'+str(i)].value=1-(1-n['T'+str(m)].value)*(n['A'+str(j)].value)/(n['A'+str(m)].value)
            s['U'+str(i)].value=n['U'+str(j+1)].value + s['T'+str(i)].value/s['S'+str(i)].value - n['T'+str(j+1)].value/n['S'+str(j+1)].value + (n['T'+str(j+1)].value-s['T'+str(i)].value)*math.log(n['S'+str(j+1)].value/s['S'+str(i)].value)/(n['S'+str(j+1)].value-s['S'+str(i)].value)
    for k in range(i-1,4,-1):
        try:
            s['T'+str(k)].value=1-(1-s['T'+str(i)].value)*(s['A'+str(k)].value-1)/(s['A'+str(i)].value-1)
        except:
            s['T'+str(k)].value=1  #gives 0/0 in the general form
        try:
            if s['S'+str(k+1)].value!=None:   
                if s['S'+str(k)].value!=s['S'+str(k+1)].value:
                    s['U'+str(k)].value=s['U'+str(k+1)].value + s['T'+str(k)].value/s['S'+str(k)].value - s['T'+str(k+1)].value/s['S'+str(k+1)].value + (s['T'+str(k+1)].value-s['T'+str(k)].value)*math.log(s['S'+str(k+1)].value/s['S'+str(k)].value)/(s['S'+str(k+1)].value-s['S'+str(k)].value)
                else:   #formula is p1=p2+y1/t1-y2/t2+(y2-y1)*ln(t2/t1)/(t2-t1)    using trapeziums for approximation (assumed percentile to be linear w.r.t. avg between data points, then integrated -y/t^2 dt from t2 to t1)
                    s['U'+str(k)].value=s['U'+str(k+1)].value  #would give 0/0 in the general form
            else:
                s['U'+str(k)].value=s['T'+str(k)].value/s['S'+str(k)].value
        except Exception as e:
            print(e)
            #s.title=s.title+' - failed'
            #break

def calculate(calcing, wbl, instr, infstr, progress):
    calcing.set(1)
    wbl.append(openpyxl.load_workbook(infstr.get(), data_only=True))
    sheets = []
    for sname in wbl[0].sheetnames:
        sheets=sheets+[wbl[0][sname]]
    del sheets[0]
    normList=['333-1','333oh-1','444-1','222-1','pyram-1','skewb-1','sq1-1','333ft-1','555-1','minx-1','666-1','333fm-1','777-1','clock-1','333bf-1','444bf-1','333mbf-1','555bf-1']
    nSheet=sheets[0]
    for sname in normList:
        if sname in sheets:
            nSheet=wbl[0][sname]
            sheets.insert(0,sheets.pop(sheets.index(sname)))
    for sheet in sheets:
        calc_t(sheet)
        if sheet.title[-1]=='1':
            calc_yp(sheet,nSheet)
        else:
            calc_yp(sheet,wbl[0][sheet.title[:-1]+'1'])
    sheet=wbl[0]['Registration']
    e=7
    while sheet[4][e].value!=None:
        e=e+1
    m=e
    for i in range(4,1+sheet.max_row):
        for e in range(7,m):
            if sheet[i][e].value==1:
                sheet[i][e].value=0
                for j in range(1,5):
                    if str(sheet[3][e].value)+'-'+str(j) in wbl[0].sheetnames:
                        nSheet=wbl[0][str(sheet[3][e].value)+'-'+str(j)]
                    else:
                        break
                    k=5
                    while k<=nSheet.max_row:
                        if nSheet['B'+str(k)].value==sheet['B'+str(i)].value:
                            break
                        k=k+1
                    if nSheet['U'+str(k)].value!=None:
                        sheet[i][e].value=(sheet[i][e].value*(j-1)/j)+(nSheet['U'+str(k)].value/j)
                    else:
                        break
            elif sheet[i][e].value==None:
                break
        progress.set(100*(i-4)/(sheet.max_row-4))
    sheet['G3'].value='POINTS'
    for i in range(4,1+sheet.max_row):
        sheet['G'+str(i)].value=0
        for e in range(7,m):
            sheet['G'+str(i)].value=(sheet['G'+str(i)].value*(e-7)+sheet[i][e].value)/(e-6)
        sheet['G'+str(i)].value=sheet['G'+str(i)].value*100
    infstr.set("Points calculated")
    calcing.set(0)
    
def showRes(wb, x, rankstr, partistr, pointstr):
    topX = []
    rankstr.set("")
    partistr.set("")
    pointstr.set("")
    sheet = wb['Registration']
    sheet['G1'].value = 0
    for i in range(0,x):
        m = 1
        for j in range(4,sheet.max_row):
            if sheet['G'+str(j)].value>sheet['G'+str(m)].value and sheet['B'+str(j)].value not in topX:
                m = j
        topX = topX+[sheet['B'+str(m)].value,sheet['G'+str(m)].value]
    for i in range(0,x):
        rankstr.set(rankstr.get()+str(i+1)+"\n")
        partistr.set(partistr.get()+topX[2*i]+"\n")
        pointstr.set(pointstr.get()+str(topX[2*i+1])+"\n")
        # print(str(i+1)+'. '+topX[2*i]+' - '+str(topX[2*i+1]))
    # print('points compiled, do you want to save it? (y/n)')
    # if input()=='y':
    #     print('enter save name for points file (.xlsx will be added automatically): ')
    #     fName=input()
    #     wb.save(fName+'.xlsx')
    #     print('done')
    # print('Press Enter to quit')
    # input()